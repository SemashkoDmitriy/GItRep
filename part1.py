from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
import time

link = "https://gknives.ru/searchingresults/manufacturer/air.html"
# загружаем книгу
file = 'links.xlsx'
links = load_workbook(file)
# делаем единственный лист активным
ws = links['Links']
#Меняем название активного листа
#ws.title = "Links"


try:
    browser = webdriver.Chrome()
    browser.get(link)
    i:int = 2 #номер строки
    st:str #адрес ячейки
    linkss = browser.find_elements(By.CSS_SELECTOR, ".image a")
    for j in range(1,100):
        for link in linkss:
            href = link.get_attribute('href')
            st = 'A'+str(i)
            ws[st] = href
            i += 1
            print(href)
        next_page = browser.find_element(By.CSS_SELECTOR, "[aria-label = 'Go to  page']")
        next_page.click()
        linkss = browser.find_elements(By.CSS_SELECTOR, ".image a")
        print(len(linkss))

except NoSuchElementException:
    links.save(file)
    # успеваем скопировать код за 30 секунд
    time.sleep(3)
    # закрываем браузер после всех манипуляций
    browser.quit()